import openpyxl
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import numpy as np

def perform_regression_and_prediction(input_path, file_name, prediction_steps):
    try:
        # 构建文件路径
        file_path = f'{input_path}{file_name}.xlsx'

        # 读取 Excel 文件
        wb = load_workbook(file_path)
        ws = wb.active

        # 提取指定列的数据
        data_column = 'D'
        column_index = ws[data_column][1].column
        data = [cell.value for cell in ws[data_column][2:]]
        index = [i for i in range(2, len(data) + 2)]

        # 创建线性回归模型
        model = LinearRegression()
        model.fit([[i] for i in index], [[d] for d in data])

        # 进行均值回归
        regressed_data = model.predict([[i] for i in index])

        # 将结果写入新工作表
        ws_new = wb.create_sheet(title='RGSTOMEAN')
        ws_new.append(['Index', data_column, 'Regressed'])

        for i, d, reg_d in zip(index, data, regressed_data):
            ws_new.append([i, d, reg_d[0]])

        # 设置表头字体加粗
        for cell in ws_new[1]:
            cell.font = Font(bold=True)

        # 预测接下来的60项数据
        last_index = index[-1]
        for i in range(last_index + 1, last_index + prediction_steps + 1):
            predicted_value = model.predict([[i]])[0][0]
            ws_new.append([i, None, predicted_value])

        # 保存结果
        wb.save(file_path)
        print(f'均值回归结果和接下来的{prediction_steps}项预测数据已保存到工作表 RGSTOMEAN')

    except FileNotFoundError:
        print(f'找不到文件 "{file_path}"，请确保路径和文件名正确并存在。')

    except Exception as e:
        print(f'发生错误: {e}')

# 用户输入文件名（不含后缀）
#file_name = input("请输入文件名（不含后缀, 但文件格式必须是.xlsx）：") # 为了在main中调用此函数, 注释掉了该行代码, 单独运行则需要去掉此行前面的'#'号

# 调用函数，传入文件路径和文件名
#input_path = 'C:\\Users\\ThinkPad\\SynologyDrive\\Trademe\\' # 为了在main中调用此函数, 注释掉了该行代码, 单独运行则需要去掉此行前面的'#'号 
#perform_regression_and_prediction(input_path, file_name) # 为了在main中调用此函数, 注释掉了该行代码, 单独运行则需要去掉此行前面的'#'号

# 文件路径为C:\\Users\\ThinkPad\\SynologyDrive\\Trademe\\
# 输入文件名(不含后缀, 但后缀必须是.xlsx)
# 数据必须在D列, 如果不是, 可以改代码
# 程序功能是对数据进行均值回归统计, 并预测接下来60项, 结果写在新的工作表RGSTOMEAN
# 程序默认预测接下来60项, 可以对代码进行修改, 以预测其他数量的项
